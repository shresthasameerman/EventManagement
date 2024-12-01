import sqlite3
import openpyxl
from openpyxl import Workbook
import os

def initialize_excel():
    if not os.path.exists('events.xlsx'):
        wb = Workbook()
        # Event Details Sheet
        ws = wb.active
        ws.title = "EventDetails"
        ws.append(["Event Name", "Event ID", "Event Date", "Event Time", "Event Duration"])
        
        # Ticket Details Sheet
        ws = wb.create_sheet("TicketDetails")
        ws.append(["Customer Name", "Ticket ID", "Event Name", "Event ID", "Event Date", "Event Time", "Duration"])
        
        wb.save("events.xlsx")

def EventDetails():
    initialize_excel()
    wb = openpyxl.load_workbook("events.xlsx")
    ws = wb["EventDetails"]
    event_details = list(ws.iter_rows(values_only=True))[1:]  # Skip headers
    
    event_names, event_ids, event_dates, event_times, event_durations = [], [], [], [], []
    for row in event_details:
        event_names.append(row[0])
        event_ids.append(row[1])
        event_dates.append(row[2])
        event_times.append(row[3])
        event_durations.append(row[4])
    
    return event_names, event_ids, event_dates, event_times, event_durations, event_details

def TicketDetails():
    conn = sqlite3.connect('event_database.db')
    cursor = conn.cursor()

    cursor.execute("CREATE TABLE IF NOT EXISTS ticket_details (customer_name TEXT, ticket_id TEXT PRIMARY KEY, event_name TEXT, event_id TEXT, event_date TEXT, event_time TEXT, duration Text)")

    cursor.execute('SELECT * FROM ticket_details')
    ticket_details = cursor.fetchall()
    
    customer_names = []
    ticket_ids = []
    event_names = []
    event_ids = []
    event_dates = []
    event_times = []
    event_durations = []
        
    conn.close()
    
    for i in ticket_details:
        customer_names.append(i[0])
        ticket_ids.append(i[1])
        event_names.append(i[2])
        event_ids.append(i[3])
        event_dates.append(i[4])
        event_times.append(i[5])
        event_durations.append(i[6])
        
    return customer_names, ticket_ids, event_names, event_ids, event_dates, event_times, event_durations, ticket_details

def BookTicket(customer_name, ticket_id, event_name):
    try:
        event_names, event_ids, event_dates, event_times, event_durations = EventDetails()[:5]
        if event_name not in event_names:
            return "Event not found."
        
        index = event_names.index(event_name)
        event_id, event_date, event_time, event_duration = event_ids[index], event_dates[index], event_times[index], event_durations[index]
        
        wb = openpyxl.load_workbook("events.xlsx")
        ws = wb["TicketDetails"]
        ws.append([customer_name, ticket_id, event_name, event_id, event_date, event_time, event_duration])
        wb.save("events.xlsx")
        return "Success"
    except Exception as e:
        return str(e)

# Create a New Event
def CreateNewEvent(event_name, event_id, event_date, event_time, event_duration):
    try:
        wb = openpyxl.load_workbook("events.xlsx")
        ws = wb["EventDetails"]
        ws.append([event_name, event_id, event_date, event_time, event_duration])
        wb.save("events.xlsx")
        return "Success"
    except Exception as e:
        return str(e)
        
def DeleteTicket(ticket_id):
    try:
        wb = openpyxl.load_workbook("events.xlsx")
        ws = wb["TicketDetails"]
        rows = list(ws.iter_rows(values_only=True))
        ws.delete_rows(1, ws.max_row)
        
        for row in rows:
            if row[1] != ticket_id:  # Keep all rows except the one to delete
                ws.append(row)
        
        wb.save("events.xlsx")
        return "Success"
    except Exception as e:
        return str(e)